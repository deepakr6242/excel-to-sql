import logging

#third party
from openpyxl import load_workbook
import sqlite3

#logger for stack trace
FORMAT = '%(asctime)-15s   %(message)s'
logging.basicConfig(level = logging.DEBUG, format = FORMAT)



#Loading the Exel sheet from abs path considering  as local machine
try:
		wb=load_workbook("D:\\oeg\\data_updated.xlsx",data_only=True)
		ws = wb.get_sheet_by_name("US & Sum")

except IOError  as e:
	logging.info('Erro due to {}'.format(e))
	raise  SystemExit("SystemExited")
		
except KeyError as k:
		logging.info('Erro due to {}'.format(k))
		SystemExit("SystemExited")


conn =  sqlite3.connect("D:\\oeg\\oeg.db")
cur = conn.cursor()



class xl_sql():

	""" get the filtered dat from excel"""

#data from Excel #
	def filtered_data(self,x):

		ndid_dict = {"Artesia": 45,"Corpus Christi ": 600,"Denver":2,"Ft. Worth":14,"Midland":10,"Oklahoma City":10,"San Antonio" :1,
		 			"United States": 1000,"International": 2000,"Other":293,"Consolidated":100}
		
		dvid_dict = {"Artesia": 45,"Corpus Christi ": 600,"Denver":2,"Ft. Worth":14,"Midland":10,"Oklahoma City":10,"San Antonio" :1,
		 			"United States": 1000,"International": 2000,"Other":293,"Consolidated":100}

	 	
	 	for i in range(1,400):
	 		if ws.cell(i,1).value == x: 
	 			# print "{} is   found ".format(x)
				row_nu = ws.cell(i,1)
				# print "row_number  of MMCFD is {}".format(row_nu)
				# print "running fo region_name {}".format(x)
				
				for r in range(i+2,i+12):
					temp = []
					
			 		# region_name = ws.cell(row =r, column = 1).value
			 		temp.append(x)
			 		temp.append(ws.cell(row =r, column = 1).value)
			 		temp.append(ndid_dict[ws.cell(row =r, column = 1).value])
			 		temp.append( dvid_dict[ws.cell(row =r, column = 1).value])
			 		temp.append(ws.cell(row =r, column = 4).value)
			 		temp.append( ws.cell(row =r, column = 5).value)
			 		temp.append( ws.cell(row =r, column = 6).value)
			 		temp.append( ws.cell(row =r, column = 7).value)
			 		temp.append( ws.cell(row =r, column = 8).value)
			 		temp.append( ws.cell(row =r, column = 9).value)
			 		temp.append( ws.cell(row =r, column = 10).value)
			 		temp.append(ws.cell(row =r, column = 11).value)
			 		temp.append(ws.cell(row =r, column = 12).value)
			 		temp.append(ws.cell(row =r, column = 13).value)
			 		temp.append( ws.cell(row =r, column = 14).value)
			 		temp.append( ws.cell(row =r, column = 15).value)		 		
			 		# print "region  is" ,region_name
			 		# print "node_id is" ,node_id
			 		# print "division_id is ",division_id
			 		# print "jan_data is ",jan_data
			 		# print "feb_data is ",feb_data
			 		# print "mar_data is ",mar_data
			 		# print "apr_data is ",apr_data
			 		# print "may_data is ",may_data
			 		# print "jun data is ",jun_data
			 		# print "july data is ",jul_data
			 		# print "August data is ",aug_data
			 		# print "sept data is ",sep_data
			 		# print "oct data is ",oct_data
			 		# print "nov data is ",nov_data
			 		# print "dec data is ",dec_data
		 			# print "--------------------------"
					# print "temp is ",temp
					to_sql(temp)

# sending the filtered  data as argument to


def  to_sql(d):

	query = ''' INSERT INTO oeg(product,product_name,node_id,division_id,jan,feb,march,april,may,june,
									july,august,sept,oct,nov,dec)
              		VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) '''
	try:    		
		cur.executemany(query, [d])                    
		conn.commit()
		# print "Executed"
	except Exception as e:
		logging.info('Erro due to {}'.format(k))
		SystemExit("SystemExited")






if __name__ == '__main__':

	data_name = ["Natural Gas Volumes (MMCFD):","CRUDE & COND. VOLUMES (MBD)","MMCFED Volumes"]
	instance = xl_sql()

	try:
		for x in data_name:
	 		instance.filtered_data(x)
	except Exception as e:
		logging.info('Erro due to  {} '.format(e))